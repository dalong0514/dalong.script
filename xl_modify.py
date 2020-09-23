# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl import load_workbook

filename = 'mdata.xlsx'
new_filename = 'mdata2.xlsx'

# wb = Workbook()
wb = load_workbook(filename)

# grab the active worksheet
# ws = wb.active
ws = wb['综合材料表']

ws['C58'] = '大龙'

# save the file
wb.template = False
wb.save(new_filename,as_template=False)